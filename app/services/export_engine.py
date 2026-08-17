"""Async XLSX export engine (Sprint 2, queue_export).

Generates ERP-identical XLSX: title row, company name, meta block (Generated /
Period / Branch / Mode / Generated Username / Report File Name), header row,
data rows. Uploads to Supabase Storage bucket `report-exports` (private);
signed URLs (24h) are created on demand by the API.
"""
import io
import json
import os
import typing
from datetime import datetime

import httpx
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.core.worker import celery_app
from app.core.db import get_db_connection
from app.services.reports import REPORTS, iter_report_rows

SUPABASE_URL = os.getenv("SUPABASE_URL", "").rstrip("/")
SUPABASE_SECRET_KEY = os.getenv("SUPABASE_SECRET_KEY", "")
BUCKET = "report-exports"


def _company_name(cur, company_id: int) -> str:
    cur.execute("SELECT company_name FROM company_configs WHERE id = %s", (company_id,))
    row = cur.fetchone()
    return row["company_name"] if row else str(company_id)


def _build_xlsx(title: str, company_name: str, meta: typing.List[typing.Tuple[str, str]],
                columns: list, rows: typing.Iterable[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    ws.cell(row=1, column=1, value=title).font = ws.cell(row=1, column=1).font.copy(bold=True)
    ws.cell(row=2, column=1, value=company_name)

    r = 4
    for label, value in meta:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=value)
        r += 1
    r += 1  # blank line like ERP

    header_row = r
    for c, col in enumerate(columns, start=1):
        cell = ws.cell(row=r, column=c, value=col["label"])
        cell.font = cell.font.copy(bold=True)
    r += 1

    for row in rows:
        for c, col in enumerate(columns, start=1):
            ws.cell(row=r, column=c, value=row.get(col["key"]))
        r += 1

    # reasonable column widths
    for c, col in enumerate(columns, start=1):
        width = max(12, min(32, len(col["label"]) + 4))
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _sb_headers(content_type: typing.Optional[str] = None) -> dict:
    """New-style sb_secret keys require the apikey header for Storage REST."""
    h = {"apikey": SUPABASE_SECRET_KEY, "Authorization": f"Bearer {SUPABASE_SECRET_KEY}"}
    if content_type:
        h["Content-Type"] = content_type
    return h


def _storage_upload(file_name: str, data: io.BytesIO) -> str:
    """Upload via Supabase Storage REST (service role). Returns storage path."""
    url = f"{SUPABASE_URL}/storage/v1/object/{BUCKET}/{file_name}"
    resp = httpx.post(
        url,
        headers={**_sb_headers("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                 "x-upsert": "true"},
        content=data.getvalue(),
        timeout=300.0,
    )
    if resp.status_code not in (200, 201):
        raise RuntimeError(f"Storage upload failed {resp.status_code}: {resp.text[:200]}")
    return file_name


def ensure_bucket():
    """Create the private export bucket once (idempotent)."""
    if not (SUPABASE_URL and SUPABASE_SECRET_KEY):
        raise RuntimeError("SUPABASE_URL / SUPABASE_SECRET_KEY not configured")
    resp = httpx.post(
        f"{SUPABASE_URL}/storage/v1/bucket",
        headers=_sb_headers("application/json"),
        json={"id": BUCKET, "name": BUCKET, "public": False},
        timeout=30.0,
    )
    # 200/201 = created; 409 (HTTP or body) already exists = fine.
    # Supabase sometimes returns HTTP 400 with a 409 "BucketAlreadyExists" body.
    body = resp.text or ""
    already = (resp.status_code == 409 or "BucketAlreadyExists" in body
               or "already exists" in body.lower())
    if resp.status_code not in (200, 201) and not already:
        raise RuntimeError(f"Bucket create failed: {resp.text[:200]}")


def make_signed_url(file_path: str, hours: int = 24) -> str:
    resp = httpx.post(
        f"{SUPABASE_URL}/storage/v1/object/sign/{BUCKET}/{file_path}",
        headers=_sb_headers("application/json"),
        json={"expiresIn": hours * 3600},
        timeout=30.0,
    )
    resp.raise_for_status()
    return f"{SUPABASE_URL}/storage/v1{resp.json()['signedURL']}"


@celery_app.task(name="app.services.export_engine.generate_export", bind=True)
def generate_export(self, export_id: str):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT company_id, requested_by, report_slug, params FROM export_files WHERE id = %s",
                    (export_id,))
        job = cur.fetchone()
        if not job:
            return f"export {export_id} not found"

        cur.execute("UPDATE export_files SET status='running' WHERE id=%s", (export_id,))
        conn.commit()

        params = job["params"] or {}
        from datetime import date
        date_from = date.fromisoformat(params["date_from"])
        date_to = date.fromisoformat(params["date_to"])
        branch = params.get("branch_esb_id") or None

        spec = REPORTS.get(job["report_slug"])
        if not spec:
            raise ValueError(f"Unknown report {job['report_slug']}")
        company_name = _company_name(cur, job["company_id"])
        now = datetime.now()
        stamp = now.strftime("%Y%m%d%H%M%S")
        file_name = f"{spec['title']} - {stamp}.xlsx"

        meta = [
            ("Generated", now.strftime("%d-%m-%Y %H:%M:%S")),
            ("Period", f"{date_from.strftime('%d-%m-%Y')} - {date_to.strftime('%d-%m-%Y')}"),
            ("Branch", params.get("branch_label") or "All"),
            ("Generated Username", job["requested_by"] or "system"),
            ("Report File Name", file_name.replace(".xlsx", "")),
        ]

        def rows_iter():
            for row in iter_report_rows(job["report_slug"], job["company_id"], branch, date_from, date_to):
                yield row

        buf = _build_xlsx(spec["title"], company_name, meta, spec["columns"], rows_iter())
        ensure_bucket()
        path = _storage_upload(f"{job['company_id']}/{file_name}", buf)

        cur.execute("""
            UPDATE export_files SET status='ready', file_path=%s, completed_at=NOW() WHERE id=%s
        """, (path, export_id))
        conn.commit()
        return f"export ready: {path}"
    except Exception as e:
        try:
            cur.execute("UPDATE export_files SET status='failed', error_message=%s WHERE id=%s",
                        (str(e)[:500], export_id))
            conn.commit()
        except Exception:
            pass
        raise
    finally:
        cur.close()
        conn.close()
